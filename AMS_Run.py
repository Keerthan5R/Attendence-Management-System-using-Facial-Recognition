"""
Attendance Management System using Facial Recognition
======================================================
Author: Keerthan R
Description:
    A Tkinter-based desktop application for managing student attendance
    using real-time facial recognition (OpenCV LBPH).
    Features:
      - Register new students by capturing face images
      - Train the face recognition model
      - Auto-mark attendance via facial recognition
      - Manually fill attendance
      - Export attendance to Excel
      - Email notification on attendance marked
"""

import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import csv
import os
import numpy as np
from PIL import Image, ImageTk
import pandas as pd
import datetime
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── Path Constants (relative — works on any machine) ───────────────────────
BASE_DIR             = os.path.dirname(os.path.abspath(__file__))
ATTENDANCE_DIR       = os.path.join(BASE_DIR, "Attendance")
STUDENT_DETAILS_DIR  = os.path.join(BASE_DIR, "StudentDetails")
TRAINING_IMAGE_DIR   = os.path.join(BASE_DIR, "TrainingImage")
TRAINING_LABEL_DIR   = os.path.join(BASE_DIR, "TrainingImageLabel")
HAARCASCADE_PATH     = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")
TRAINER_PATH         = os.path.join(TRAINING_LABEL_DIR, "trainer.yml")
STUDENT_CSV          = os.path.join(STUDENT_DETAILS_DIR, "StudentDetails.csv")

# Create required folders if they don't exist
for folder in [ATTENDANCE_DIR, STUDENT_DETAILS_DIR, TRAINING_IMAGE_DIR, TRAINING_LABEL_DIR]:
    os.makedirs(folder, exist_ok=True)

# ─── Main Window Setup ───────────────────────────────────────────────────────
window = tk.Tk()
window.title("Attendance Management System — Facial Recognition")
window.geometry("1280x720")
window.configure(background="#1e1e2e")
window.resizable(True, True)

# ─── Color Palette ───────────────────────────────────────────────────────────
BG_DARK   = "#1e1e2e"
BG_PANEL  = "#2a2a3e"
ACCENT    = "#7c3aed"
ACCENT2   = "#06b6d4"
TEXT_MAIN = "#f8f8f2"
TEXT_SUB  = "#a8a8c0"
SUCCESS   = "#22c55e"
WARNING   = "#f59e0b"
DANGER    = "#ef4444"

# ─── Fonts ───────────────────────────────────────────────────────────────────
FONT_TITLE  = ("Helvetica", 22, "bold")
FONT_HEAD   = ("Helvetica", 14, "bold")
FONT_LABEL  = ("Helvetica", 12)
FONT_ENTRY  = ("Helvetica", 13)
FONT_BTN    = ("Helvetica", 11, "bold")

# ─── Helper: Styled Button ────────────────────────────────────────────────────
def make_btn(parent, text, cmd, color=ACCENT, fg=TEXT_MAIN, width=18):
    btn = tk.Button(
        parent, text=text, command=cmd,
        bg=color, fg=fg, font=FONT_BTN,
        relief="flat", bd=0, width=width,
        activebackground=ACCENT2, activeforeground=TEXT_MAIN,
        cursor="hand2", pady=8
    )
    btn.bind("<Enter>", lambda e: btn.config(bg=ACCENT2))
    btn.bind("<Leave>", lambda e: btn.config(bg=color))
    return btn

# ─── Helper: Show Toast Notification ─────────────────────────────────────────
def show_toast(message, color=SUCCESS, duration=3000):
    toast = tk.Toplevel(window)
    toast.overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.configure(bg=color)
    toast.geometry(f"420x50+{window.winfo_x()+430}+{window.winfo_y()+680}")
    tk.Label(toast, text=message, bg=color, fg="white",
             font=FONT_LABEL, pady=10, padx=20).pack(fill="both")
    toast.after(duration, toast.destroy)

# ─── Helper: Export DataFrame to styled Excel ────────────────────────────────
def export_to_excel(df, filepath, sheet_name="Attendance"):
    """Export a pandas DataFrame to a styled Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    alt_fill    = PatternFill("solid", fgColor="2A2A3E")

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filepath)

# ─── Section: Take Images ────────────────────────────────────────────────────
def take_images():
    """Open a sub-window to capture student face images."""
    win = tk.Toplevel(window)
    win.title("Register Student — Capture Images")
    win.geometry("600x380")
    win.configure(bg=BG_PANEL)
    win.grab_set()

    tk.Label(win, text="📷  Register New Student", font=FONT_TITLE,
             bg=BG_PANEL, fg=TEXT_MAIN).pack(pady=20)

    frame = tk.Frame(win, bg=BG_PANEL)
    frame.pack(pady=10)

    tk.Label(frame, text="Enrollment No:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).grid(row=0, column=0, sticky="w", padx=10, pady=8)
    enr_entry = tk.Entry(frame, font=FONT_ENTRY, width=25,
                         bg="#3a3a5c", fg=TEXT_MAIN, insertbackground=TEXT_MAIN)
    enr_entry.grid(row=0, column=1, padx=10)

    tk.Label(frame, text="Student Name:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).grid(row=1, column=0, sticky="w", padx=10, pady=8)
    name_entry = tk.Entry(frame, font=FONT_ENTRY, width=25,
                          bg="#3a3a5c", fg=TEXT_MAIN, insertbackground=TEXT_MAIN)
    name_entry.grid(row=1, column=1, padx=10)

    status_lbl = tk.Label(win, text="", font=FONT_LABEL, bg=BG_PANEL, fg=SUCCESS)
    status_lbl.pack(pady=8)

    def capture():
        enrollment = enr_entry.get().strip()
        name       = name_entry.get().strip()

        if not enrollment.isdigit():
            messagebox.showwarning("Invalid Input", "Enrollment must be a number.", parent=win)
            return
        if not name:
            messagebox.showwarning("Invalid Input", "Student name cannot be empty.", parent=win)
            return

        # Save student details to CSV
        row = [enrollment, name]
        if not os.path.exists(STUDENT_CSV):
            with open(STUDENT_CSV, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Enrollment", "Name"])
                writer.writerow(row)
        else:
            with open(STUDENT_CSV, "a", newline="") as f:
                csv.writer(f).writerow(row)

        cam = cv2.VideoCapture(0)
        detector = cv2.CascadeClassifier(HAARCASCADE_PATH)
        count = 0
        status_lbl.config(text="📸 Capturing... Please look at the camera.")
        win.update()

        while count < 100:
            ret, frame = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
            for (x, y, w, h) in faces:
                count += 1
                img_path = os.path.join(
                    TRAINING_IMAGE_DIR, f"{name}.{enrollment}.{count}.jpg"
                )
                cv2.imwrite(img_path, gray[y:y+h, x:x+w])
                cv2.rectangle(frame, (x, y), (x+w, y+h), (127, 0, 255), 2)
                cv2.putText(frame, f"Captured: {count}/100", (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.imshow("Capturing Faces — Press Q to stop early", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break

        cam.release()
        cv2.destroyAllWindows()
        status_lbl.config(text=f"✅ {count} images captured for {name} (ID: {enrollment})")
        show_toast(f"✅ {count} images saved for {name}!")

    btn_frame = tk.Frame(win, bg=BG_PANEL)
    btn_frame.pack(pady=15)
    make_btn(btn_frame, "📸  Capture Images", capture).pack(side="left", padx=10)
    make_btn(btn_frame, "✖  Close", win.destroy, color="#4a4a6a").pack(side="left", padx=10)

# ─── Section: Train Model ─────────────────────────────────────────────────────
def train_images():
    """Train the LBPH face recognizer on captured images."""
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    detector   = cv2.CascadeClassifier(HAARCASCADE_PATH)

    faces, ids = [], []
    image_paths = [
        os.path.join(TRAINING_IMAGE_DIR, f)
        for f in os.listdir(TRAINING_IMAGE_DIR)
        if f.lower().endswith((".jpg", ".jpeg", ".png"))
    ]

    if not image_paths:
        messagebox.showwarning("No Images", "No training images found! Please capture images first.")
        return

    for img_path in image_paths:
        try:
            pil_img = Image.open(img_path).convert("L")
            img_arr = np.array(pil_img, dtype=np.uint8)
            # Parse enrollment ID from filename: Name.EnrollmentID.Count.jpg
            filename  = os.path.basename(img_path)
            parts     = filename.split(".")
            student_id = int(parts[1]) if len(parts) >= 3 else 0
            detected   = detector.detectMultiScale(img_arr)
            for (x, y, w, h) in detected:
                faces.append(img_arr[y:y+h, x:x+w])
                ids.append(student_id)
        except Exception as e:
            print(f"[WARN] Skipping {img_path}: {e}")

    if not faces:
        messagebox.showwarning("Training Failed", "No faces detected in training images.")
        return

    recognizer.train(faces, np.array(ids))
    recognizer.write(TRAINER_PATH)

    show_toast(f"✅ Model trained on {len(faces)} face samples!")
    messagebox.showinfo("Training Complete",
                        f"Model trained on {len(faces)} samples.\nSaved to: {TRAINER_PATH}")

# ─── Section: Auto Attendance (Facial Recognition) ───────────────────────────
def auto_attendance():
    """Mark attendance automatically using facial recognition."""
    if not os.path.exists(TRAINER_PATH):
        messagebox.showerror("Model Not Found",
                             "No trained model found!\nPlease train the model first.")
        return

    win = tk.Toplevel(window)
    win.title("Auto Attendance — Facial Recognition")
    win.geometry("420x180")
    win.configure(bg=BG_PANEL)
    win.grab_set()

    tk.Label(win, text="🎯  Auto Attendance", font=FONT_TITLE,
             bg=BG_PANEL, fg=TEXT_MAIN).pack(pady=15)
    tk.Label(win, text="Subject Name:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).pack()
    sub_entry = tk.Entry(win, font=FONT_ENTRY, width=28,
                         bg="#3a3a5c", fg=TEXT_MAIN, insertbackground=TEXT_MAIN)
    sub_entry.pack(pady=6)
    status_lbl = tk.Label(win, text="", font=FONT_LABEL, bg=BG_PANEL, fg=SUCCESS)
    status_lbl.pack()

    def run_recognition():
        subject = sub_entry.get().strip()
        if not subject:
            messagebox.showwarning("Missing Subject", "Please enter subject name.", parent=win)
            return

        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read(TRAINER_PATH)
        detector = cv2.CascadeClassifier(HAARCASCADE_PATH)

        # Load student details
        student_map = {}
        if os.path.exists(STUDENT_CSV):
            df = pd.read_csv(STUDENT_CSV)
            for _, row in df.iterrows():
                student_map[int(row["Enrollment"])] = row["Name"]

        now       = datetime.datetime.now()
        date_str  = now.strftime("%Y-%m-%d")
        time_str  = now.strftime("%H-%M-%S")
        xlsx_name = f"{subject}_{date_str}_{time_str}.xlsx"
        xlsx_path = os.path.join(ATTENDANCE_DIR, xlsx_name)

        attendance_records = []
        marked_ids = set()

        cam = cv2.VideoCapture(0)
        status_lbl.config(text="📹 Running... Press Q to stop.")
        win.update()

        while True:
            ret, frame = cam.read()
            if not ret:
                break
            gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

            for (x, y, w, h) in faces:
                student_id, confidence = recognizer.predict(gray[y:y+h, x:x+w])
                label = "Unknown"
                color = (0, 0, 255)

                if confidence < 70 and student_id in student_map:
                    label = f"{student_map[student_id]} ({student_id})"
                    color = (0, 255, 0)
                    if student_id not in marked_ids:
                        marked_ids.add(student_id)
                        attendance_records.append({
                            "Enrollment": student_id,
                            "Name": student_map[student_id],
                            "Subject": subject,
                            "Date": date_str,
                            "Time": now.strftime("%H:%M:%S"),
                            "Status": "Present"
                        })
                        # Send email notification (non-blocking)
                        try:
                            from notifier import send_attendance_email
                            send_attendance_email(
                                student_name=student_map[student_id],
                                subject=subject,
                                date=date_str,
                                time_val=now.strftime("%H:%M:%S")
                            )
                        except Exception:
                            pass  # Email is optional

                cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                cv2.putText(frame, label, (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
                cv2.putText(frame, f"Conf: {confidence:.1f}", (x, y+h+20),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 0), 1)

            cv2.putText(frame, f"Marked: {len(marked_ids)} | Press Q to stop",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.imshow("Auto Attendance — Facial Recognition", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break

        cam.release()
        cv2.destroyAllWindows()

        if attendance_records:
            df_out = pd.DataFrame(attendance_records)
            export_to_excel(df_out, xlsx_path)
            status_lbl.config(text=f"✅ {len(marked_ids)} students marked. Saved!")
            show_toast(f"✅ Attendance saved: {xlsx_name}")
            messagebox.showinfo("Attendance Saved",
                                f"Marked {len(marked_ids)} students.\nFile: {xlsx_path}", parent=win)
        else:
            status_lbl.config(text="⚠ No faces recognized.")
            show_toast("⚠ No faces recognized.", color=WARNING)

    btn_frame = tk.Frame(win, bg=BG_PANEL)
    btn_frame.pack(pady=8)
    make_btn(btn_frame, "▶  Start Recognition", run_recognition).pack(side="left", padx=8)
    make_btn(btn_frame, "✖  Close", win.destroy, color="#4a4a6a").pack(side="left", padx=8)

# ─── Section: Manual Attendance ───────────────────────────────────────────────
def manually_fill():
    """Manually enter attendance for a subject."""
    win = tk.Toplevel(window)
    win.title("Manual Attendance")
    win.geometry("700x500")
    win.configure(bg=BG_PANEL)
    win.grab_set()

    tk.Label(win, text="✏️  Manual Attendance", font=FONT_TITLE,
             bg=BG_PANEL, fg=TEXT_MAIN).pack(pady=15)

    top_frame = tk.Frame(win, bg=BG_PANEL)
    top_frame.pack(fill="x", padx=20)
    tk.Label(top_frame, text="Subject:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).pack(side="left")
    sub_entry = tk.Entry(top_frame, font=FONT_ENTRY, width=22,
                         bg="#3a3a5c", fg=TEXT_MAIN, insertbackground=TEXT_MAIN)
    sub_entry.pack(side="left", padx=10)

    # Table for entries
    cols = ("Enrollment", "Name")
    tree = ttk.Treeview(win, columns=cols, show="headings", height=10)
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=200, anchor="center")
    tree.pack(padx=20, pady=10, fill="both")

    entry_frame = tk.Frame(win, bg=BG_PANEL)
    entry_frame.pack(fill="x", padx=20)
    tk.Label(entry_frame, text="Enrollment:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).pack(side="left")
    enr_e = tk.Entry(entry_frame, font=FONT_ENTRY, width=12,
                     bg="#3a3a5c", fg=TEXT_MAIN, insertbackground=TEXT_MAIN)
    enr_e.pack(side="left", padx=6)
    tk.Label(entry_frame, text="Name:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).pack(side="left")
    name_e = tk.Entry(entry_frame, font=FONT_ENTRY, width=18,
                      bg="#3a3a5c", fg=TEXT_MAIN, insertbackground=TEXT_MAIN)
    name_e.pack(side="left", padx=6)

    records = []

    def add_row():
        enr  = enr_e.get().strip()
        name = name_e.get().strip()
        if not enr or not name:
            messagebox.showwarning("Missing Data", "Fill both Enrollment and Name.", parent=win)
            return
        records.append((enr, name))
        tree.insert("", "end", values=(enr, name))
        enr_e.delete(0, "end")
        name_e.delete(0, "end")

    def save_excel():
        subject = sub_entry.get().strip()
        if not subject:
            messagebox.showwarning("Missing Subject", "Enter subject name.", parent=win)
            return
        if not records:
            messagebox.showwarning("No Data", "Add at least one student.", parent=win)
            return
        now      = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H-%M-%S")
        xlsx_path = os.path.join(ATTENDANCE_DIR,
                                 f"Manual_{subject}_{date_str}_{time_str}.xlsx")
        df = pd.DataFrame(records, columns=["Enrollment", "Name"])
        df["Subject"] = subject
        df["Date"]    = date_str
        df["Time"]    = now.strftime("%H:%M:%S")
        df["Status"]  = "Present"
        export_to_excel(df, xlsx_path)
        show_toast(f"✅ Manual attendance saved: {os.path.basename(xlsx_path)}")
        messagebox.showinfo("Saved", f"Attendance saved:\n{xlsx_path}", parent=win)
        win.destroy()

    btn_frame = tk.Frame(win, bg=BG_PANEL)
    btn_frame.pack(pady=10)
    make_btn(btn_frame, "➕  Add Student", add_row, color=ACCENT2).pack(side="left", padx=8)
    make_btn(btn_frame, "💾  Save to Excel", save_excel).pack(side="left", padx=8)
    make_btn(btn_frame, "✖  Cancel", win.destroy, color="#4a4a6a").pack(side="left", padx=8)

# ─── Section: View Attendance ─────────────────────────────────────────────────
def view_attendance():
    """Open a window to view all attendance Excel files."""
    win = tk.Toplevel(window)
    win.title("View Attendance Records")
    win.geometry("900x550")
    win.configure(bg=BG_PANEL)
    win.grab_set()

    tk.Label(win, text="📊  Attendance Records", font=FONT_TITLE,
             bg=BG_PANEL, fg=TEXT_MAIN).pack(pady=12)

    files = [f for f in os.listdir(ATTENDANCE_DIR) if f.endswith(".xlsx")]
    if not files:
        tk.Label(win, text="No attendance files found.", font=FONT_LABEL,
                 bg=BG_PANEL, fg=TEXT_SUB).pack(pady=30)
        return

    # File selector
    sel_frame = tk.Frame(win, bg=BG_PANEL)
    sel_frame.pack(fill="x", padx=20, pady=6)
    tk.Label(sel_frame, text="Select File:", font=FONT_LABEL,
             bg=BG_PANEL, fg=TEXT_SUB).pack(side="left")
    file_var = tk.StringVar(value=files[0])
    file_menu = ttk.Combobox(sel_frame, textvariable=file_var,
                             values=files, font=FONT_LABEL, width=50)
    file_menu.pack(side="left", padx=10)

    # Treeview
    tree_frame = tk.Frame(win, bg=BG_PANEL)
    tree_frame.pack(fill="both", expand=True, padx=20, pady=6)
    tree = ttk.Treeview(tree_frame, show="headings")
    vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")

    def load_file(event=None):
        filepath = os.path.join(ATTENDANCE_DIR, file_var.get())
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=win)
            return
        tree.delete(*tree.get_children())
        tree["columns"] = list(df.columns)
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=130, anchor="center")
        for _, row in df.iterrows():
            tree.insert("", "end", values=list(row))

    file_menu.bind("<<ComboboxSelected>>", load_file)
    load_file()

    make_btn(win, "✖  Close", win.destroy, color="#4a4a6a").pack(pady=10)

# ─── Main UI Layout ───────────────────────────────────────────────────────────

# Header
header = tk.Frame(window, bg=ACCENT, height=80)
header.pack(fill="x")
tk.Label(header, text="🎓  Attendance Management System",
         font=("Helvetica", 26, "bold"), bg=ACCENT, fg=TEXT_MAIN, pady=18).pack()

# Subtitle
tk.Label(window,
         text="Powered by Facial Recognition  |  OpenCV · Tkinter · Flask",
         font=("Helvetica", 11), bg=BG_DARK, fg=TEXT_SUB).pack(pady=8)

# Button Grid
btn_container = tk.Frame(window, bg=BG_DARK)
btn_container.pack(expand=True)

buttons = [
    ("📷  Register Student",    take_images,    ACCENT),
    ("🧠  Train Model",         train_images,   "#059669"),
    ("🎯  Auto Attendance",     auto_attendance, ACCENT2),
    ("✏️  Manual Attendance",   manually_fill,  WARNING),
    ("📊  View Records",        view_attendance, "#7c3aed"),
]

for i, (label, cmd, color) in enumerate(buttons):
    row = i // 3
    col = i % 3
    btn = make_btn(btn_container, label, cmd, color=color, width=22)
    btn.grid(row=row, column=col, padx=20, pady=15, ipady=10)

# Footer
footer = tk.Frame(window, bg=BG_PANEL, height=40)
footer.pack(fill="x", side="bottom")
tk.Label(footer,
         text="Keerthan R  |  Gopalan College of Engineering  |  TechSaksham — Microsoft & SAP",
         font=("Helvetica", 9), bg=BG_PANEL, fg=TEXT_SUB, pady=12).pack()

window.mainloop()
